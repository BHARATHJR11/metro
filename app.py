import os
from flask import Flask, request, render_template, redirect, url_for, send_file, flash
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads/'
app.secret_key = 'your_secret_key'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB limit

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        file1 = request.files['file1']
        file2 = request.files['file2']
        result_name = request.form['result_name']

        if file1.filename == '' or file2.filename == '':
            flash('Please select both files.')
            return redirect(request.url)

        if result_name.strip() == '':
            flash('Please enter a result file name.')
            return redirect(request.url)

        file1_path = os.path.join(app.config['UPLOAD_FOLDER'], file1.filename)
        file2_path = os.path.join(app.config['UPLOAD_FOLDER'], file2.filename)

        file1.save(file1_path)
        file2.save(file2_path)

        result_file_path = compare_files(file1_path, file2_path, result_name)
        return redirect(url_for('result', result_file=result_file_path))

    return render_template('index.html')

def compare_files(file1, file2, result_name):
    df1 = pd.read_excel(file1, sheet_name='NCMCParkingDB')
    df2 = pd.read_excel(file2, sheet_name='NCMC-ParkingSettlement')

    grouped_df1 = df1.groupby(['Name', 'Terminal_id']).sum().reset_index()
    grouped_df2 = df2.groupby(['Merchant Name', 'Terminal ID']).sum().reset_index()

    comparison_result = pd.DataFrame()
    comparison_result['Name'] = grouped_df1['Name']
    comparison_result['Terminal_id'] = grouped_df1['Terminal_id']
    comparison_result['NCMC_SVP_Amt'] = grouped_df1.iloc[:, 7]
    comparison_result['   '] = ''
    comparison_result['    '] = ''
    comparison_result['Merchant Name'] = grouped_df2['Merchant Name']
    comparison_result['Terminal ID'] = grouped_df2['Terminal ID']
    comparison_result['Settlement Amount'] = grouped_df2.iloc[:, 7]
    comparison_result['     '] = ''
    comparison_result['NCMC_SVP_Amt'] = pd.to_numeric(comparison_result['NCMC_SVP_Amt'], errors='coerce')
    comparison_result['Settlement Amount'] = pd.to_numeric(comparison_result['Settlement Amount'], errors='coerce')
    comparison_result['Difference'] = comparison_result['NCMC_SVP_Amt'] - comparison_result['Settlement Amount']

    total_NCMC_SVP_amt_sum = comparison_result['NCMC_SVP_Amt'].sum()
    total_settlement_sum = comparison_result['Settlement Amount'].sum()
    difference_sum = comparison_result['Difference'].sum()

    output_file = os.path.join(app.config['UPLOAD_FOLDER'], f'{result_name}.xlsx')
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        grouped_df1.to_excel(writer, sheet_name='NCMCParkingDB', index=False)
        grouped_df2.to_excel(writer, sheet_name='NCMC-ParkingSettlement', index=False)
        comparison_result.to_excel(writer, sheet_name='Comparison', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Comparison']

        last_row = len(comparison_result) + 2
        worksheet.cell(row=last_row, column=8, value=total_settlement_sum)
        worksheet.cell(row=last_row, column=3, value=total_NCMC_SVP_amt_sum)
        worksheet.cell(row=last_row, column=10, value=difference_sum)

    workbook = openpyxl.load_workbook(output_file)
    sheet = workbook['Comparison']

    fill_positive = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    fill_negative = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    fill_empty = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    def apply_color(cell, fill):
        cell.fill = fill

    for row in sheet.iter_rows(min_row=2, min_col=10, max_col=10):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value > 0:
                apply_color(cell, fill_positive)
            elif isinstance(cell.value, (int, float)) and cell.value < 0:
                apply_color(cell, fill_negative)

    columns_to_check = [1, 2, 3, 6, 7, 8]
    for row in sheet.iter_rows(min_row=2):
        for col in columns_to_check:
            cell = row[col - 1]
            if cell.value is None or cell.value == "":
                apply_color(cell, fill_empty)

    workbook.save(output_file)
    return output_file

@app.route('/result')
def result():
    result_file = request.args.get('result_file')
    return render_template('result.html', result_file=result_file)

@app.route('/download/<filename>')
def download(filename):
    return send_file(os.path.join(app.config['UPLOAD_FOLDER'], filename), as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
